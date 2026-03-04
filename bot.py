import requests
import json
from fastapi import FastAPI, Request
from openpyxl import load_workbook

from config import API_URL, HEADERS, ADMIN_ID
from storage import *

app = FastAPI()


# ================= UTIL =================

def send_message(chat_id, text, keyboard=None):
    payload = {
        "chat_id": chat_id,
        "text": text,
    }

    if keyboard:
        payload["inline_keyboard"] = keyboard

    requests.post(f"{API_URL}/messages", headers=HEADERS, json=payload)


def load_clubs():
    wb = load_workbook("joined_clubs.xlsx")
    sheet = wb.active
    clubs = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        clubs.append({
            "direction": row[0],
            "name": row[1],
            "age": row[2],
            "address": row[3],
            "teacher": row[4],
            "link": row[5],
        })

    return clubs


def load_masterclasses():
    try:
        with open("masterclasses.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except:
        return []


def save_masterclasses(data):
    with open("masterclasses.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


# ================= WEBHOOK =================

@app.post("/webhook")
async def webhook(request: Request):
    data = await request.json()

    if "message" in data:
        message = data["message"]
        user_id = message["from"]["id"]
        chat_id = message["chat"]["id"]
        text = message.get("text")

        await handle_message(user_id, chat_id, text)

    if "callback_query" in data:
        callback = data["callback_query"]
        user_id = callback["from"]["id"]
        chat_id = callback["message"]["chat"]["id"]
        callback_data = callback["data"]

        await handle_callback(user_id, chat_id, callback_data)

    return {"ok": True}


# ================= HANDLERS =================

async def handle_message(user_id, chat_id, text):

    state = get_state(user_id)

    if text == "/start":
        send_message(
            chat_id,
            "Привет! Я бот MAX центра 🎨\nВыберите раздел:",
            keyboard=[
                [{"text": "🎨 Кружки", "callback_data": "clubs"}],
                [{"text": "🧩 Мастер-классы", "callback_data": "masters"}],
                [{"text": "🎉 Пакеты", "callback_data": "packages"}],
                [{"text": "✉ Поддержка", "callback_data": "support"}],
            ]
        )
        return

    # SUPPORT
    if state == "support":
        send_message(
            ADMIN_ID,
            f"✉ Новое сообщение поддержки:\n\n{text}"
        )
        send_message(chat_id, "Сообщение отправлено ✅")
        clear_state(user_id)
        return

    # Пример FSM: возраст для кружков
    if state == "club_age":
        if not text.isdigit():
            send_message(chat_id, "Введите возраст числом")
            return

        update_data(user_id, "age", int(text))
        set_state(user_id, "club_direction")

        send_message(chat_id, "Выберите направление:",
                     keyboard=[
                         [{"text": "Робототехника", "callback_data": "dir_0"}],
                         [{"text": "Программирование", "callback_data": "dir_1"}],
                     ])
        return


async def handle_callback(user_id, chat_id, data):

    if data == "support":
        set_state(user_id, "support")
        send_message(chat_id, "Напишите сообщение:")
        return

    if data == "clubs":
        set_state(user_id, "club_age")
        send_message(chat_id, "Введите возраст:")
        return

    if data == "masters":
        masters = load_masterclasses()

        if not masters:
            send_message(chat_id, "Пока нет мастер-классов")
            return

        buttons = []
        for i, m in enumerate(masters):
            buttons.append([{
                "text": m["title"],
                "callback_data": f"master_{i}"
            }])

        send_message(chat_id, "Доступные мастер-классы:",
                     keyboard=buttons)
        return

    if data.startswith("master_"):
        index = int(data.split("_")[1])
        masters = load_masterclasses()

        if index >= len(masters):
            return

        m = masters[index]

        send_message(
            chat_id,
            f"{m['title']}\n\n{m['description']}\n\n"
            f"📅 {m['date']}\n💰 {m['price']} ₽\n"
            f"👩‍🏫 {m['teacher']}\n\n{m['link']}"
        )
        return